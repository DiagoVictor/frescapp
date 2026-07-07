from flask import Blueprint, jsonify, request, send_file, Response
from ..models.order import Order
import json
from flask_bcrypt import Bcrypt
from datetime import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import locale
from flask import Response
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
from io import BytesIO
from pymongo import MongoClient
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak
import locale
from ..models.customer import Customer
from ..models.product import Product
from ..models.inventory import Inventory
from datetime import datetime, timedelta
import requests

locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
report_api = Blueprint('report', __name__)
from ..db import get_db
db = get_db()
orders_collection = db['orders']
products_collection = db['orders']

_remote_image_cache = {}

def _get_cached_image(url):
    if url not in _remote_image_cache:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        _remote_image_cache[url] = response.content
    return BytesIO(_remote_image_cache[url])
@report_api.route('/picking/<string:startDate>/<string:endDate>', methods=['GET'])
def get_picking(startDate, endDate): 
    buffer = BytesIO()
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=28.35,
        rightMargin=28.35,
        topMargin=28.35,
        bottomMargin=28.35
    )

    styles = getSampleStyleSheet()
    image_path = 'https://buyfrescapp.com/images/banner1.png'
    logo = Image(_get_cached_image(image_path), width=95, height=50)

    pdf_content = []
    orders = orders_collection.find({"delivery_date": {"$gte": startDate, "$lte": endDate }}, {
        "order_number": 1, "delivery_date": 1, "customer_name": 1, "customer_phone": 1,
        "paymentMethod": 1, "deliverySlot": 1, "open_hour": 1, "deliveryAddress": 1,
        "deliveryAddressDetails": 1, "products": 1, "deliveryCost": 1, "discount": 1
    })

    for order in orders:
        if not order:
            return jsonify({'message': 'Order not found'}), 404        

        word_wrap_style = styles["Normal"]
        word_wrap_style.wordWrap = 'CJK'

        # Datos de la orden
        order_data = [
            [logo, 'Remisión #', order['order_number'], 'Fecha', order['delivery_date']],
            ['', 'Nombre', Paragraph(order['customer_name'], word_wrap_style), 'Teléfono del Cliente', Paragraph(order['customer_phone'], word_wrap_style)],
            ['', 'Método de pago', Paragraph(order['paymentMethod'], word_wrap_style), 'Horario de entrega', Paragraph(order['deliverySlot'] + ' (' + order['open_hour'] + ')', word_wrap_style)],
            ['Dirección de entrega', Paragraph(order['deliveryAddress'], word_wrap_style),  'Detalle de entrega', Paragraph(order['deliveryAddressDetails'], word_wrap_style),'']
        ]

        order_table = Table(order_data, colWidths=[100, 100, 150, 100, 100])
        order_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (0, 2)),  # Logo 3 filas
            ('VALIGN', (0, 0), (0, 2), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 2), 'CENTER'),
            # Dirección: valor solo en columna 2
            ('SPAN', (1, 3), (1, 3)),

            # Detalle: valor ocupa columnas 4 y 5
            ('SPAN', (3, 3), (4, 3)),


            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
        ]))

        pdf_content.append(order_table)
        pdf_content.append(Spacer(1, 20))

        # Aplicar estilos de WordWrap
        word_wrap_style = styles["Normal"]
        word_wrap_style.wordWrap = 'CJK'

        products = list(order['products'])
        product_data = [
            ['SKU', 'Nombre', 'Cantidad', 'Precio Unitario', 'Total'],  # Encabezado
        ]
        for product in sorted(products, key=lambda x: x['name']):
            sku = product['sku']
            name = product['name']
            quantity = product['quantity']
            price_sale = locale.format_string('%.0f', round(product.get('price_sale'),0), grouping=True)
            total = locale.format_string('%.0f', round(float(product.get('price_sale')) * float(quantity),0), grouping=True)
            name_paragraph = Paragraph(name, word_wrap_style)
            product_row = [sku, name_paragraph, str(quantity)  + "  " + str(product.get('unit', '')), price_sale, total]
            product_data.append(product_row)

        subtotal = sum(round(float(product['quantity']) * float(product['price_sale']),0) for product in products) +  order.get('deliveryCost', 0.0)
        descuentos = float(order.get('discount', 0))
        total = subtotal - descuentos
        subtotal_formatted = locale.format_string('%.2f', subtotal, grouping=True)
        descuentos_formatted = locale.format_string('%.2f', descuentos, grouping=True)
        total_formatted = locale.format_string('%.2f', total, grouping=True)
        image_path_payment = 'https://buyfrescapp.com/images/medio_pago.png'  # URL o ruta local de la imagen del medio de pago
        payment_image = Image(_get_cached_image(image_path_payment), width=290, height=80)  # Ajustar tamaño de la imagen
        costo_domicilio = locale.format_string('%.2f', order.get('deliveryCost', 0.0), grouping=True)
        product_data.extend([[payment_image,'','','Costo domicilio', costo_domicilio],
            ['', '', '', 'Subtotal', subtotal_formatted],
            ['', '', '', 'Descuentos', descuentos_formatted],
            ['', '', '', 'Total', total_formatted]
        ])
        product_table = Table(product_data, colWidths=[120,200,80,80,80])
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#97D700')),  # Color de fondo del encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir bordes internos a las celdas
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('SPAN', (0, len(products)+1), (2, len(products)+4))
        ]))
        pdf_content.append(product_table)
        pdf_content.append(PageBreak())

    pdf.build(pdf_content)
    buffer.seek(0)
    response = Response(buffer, mimetype='application/pdf')
    response.headers['Content-Disposition'] = 'inline; filename=remisiones_{}_{}.pdf'.format(startDate,endDate)
    return response

@report_api.route('/compras/<string:date>/<string:supplier>', methods=['GET'])
def get_compras(date,supplier):
    pipeline = [
        {
            "$match": {
                "delivery_date": date
            }
        },
        {
            "$unwind": "$products"
        },
        {
            "$group": {
                "_id": "$products.sku",
                "total_quantity_ordered": {"$sum": "$products.quantity"}
            }
        },
        {
            "$lookup": {
                "from": "products",
                "localField": "_id",
                "foreignField": "sku",
                "as": "product_info"
            }
        },
        {
            "$unwind": "$product_info"
        },
        {
            "$project": {
                "_id": 0,
                "sku": "$_id",
                "name": "$product_info.name",
                "total_quantity_ordered": 1,
                "price_purchase": "$product_info.price_purchase",
                "proveedor": "$product_info.proveedor",
                "category": "$product_info.category",
                "unit": "$product_info.unit"
            }
        }
    ]
    if supplier != 'Todos':
        pipeline.append({
            "$match": {
                "proveedor": supplier
            }
        })
    pipeline.append({
        "$sort": {
            "category": 1 ,
            "name":1
        }
    })
    products = list(orders_collection.aggregate(pipeline))
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    image_path = 'https://app.buyfrescapp.com:5000/api/shared/banner1.png'
    logo = Image(_get_cached_image(image_path), width=200, height=70)
    centered_style = ParagraphStyle(
            name='Centered',
            fontSize=16,  # Tamaño de la letra aumentado a 16
            alignment=TA_CENTER,  # Centrado horizontal
            textColor=colors.white,  # Color del texto blanco
            leading=50  # Espaciado entre líneas para centrar verticalmente
        )
    pdf_content = []
    compras_paragraph = Paragraph('<font>Compras para el {}</font>'.format(date), centered_style)
    green_box = Table([[compras_paragraph]], colWidths=[250], rowHeights=[70], style=[('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#97D700'))])
    content_table = Table([
    [logo, green_box]], colWidths=[200, 250])
    pdf_content.append(content_table)
    pdf_content.append(Paragraph('<br/><br/>', styles['Normal']))
    table_width = 500
    product_data = [
        [ 'Nombre', 'Categoria', 'Cantidad', 'Precio Unitario', 'Proveedor','Proveedor Final', 'Precio Final'],  # Encabezado
    ]
    word_wrap_style = styles["Normal"]
    word_wrap_style.wordWrap = 'CJK'

    for product in products:
        sku = product['sku']
        name = product['name'] + " - ( "+sku+" )"
        quantity = Paragraph(str(product.get('total_quantity_ordered')) + "  " + str(product.get('unit')),word_wrap_style)
        price = locale.format_string('%.2f', round(product.get('price_purchase'),0), grouping=True)
        proveedor = product['proveedor']
        name_paragraph = Paragraph(name, word_wrap_style)
        product_row = [ name_paragraph, product.get('category'), quantity, price, proveedor]
        product_data.append(product_row)
    total = locale.format_string('%.2f',sum(round(float(product['total_quantity_ordered']) * float(product['price_purchase']),0) for product in products), grouping=True)
    product_data.extend([['','','','',''],
            ['', '', 'Total', total, '']
        ])
    col_widths = [(2 / 8) * table_width] + [(1 / 8) * table_width] * 4
    product_table = Table(product_data, colWidths=col_widths)
    product_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#97D700')),  # Color de fondo del encabezado
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Añadir bordes internos a las celdas
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black)  # Añadir borde alrededor de la tabla
    ]))
    pdf_content.append(product_table)
    pdf_content.append(PageBreak())

    pdf.build(pdf_content)
    buffer.seek(0)
    response = Response(buffer, mimetype='application/pdf')
    response.headers['Content-Disposition'] = 'inline; filename=ordenes_{}.pdf'.format(date)
    return response

@report_api.route('/picking_summary/<string:forecastDate>', methods=['GET'])
def get_picking_summary(forecastDate):
    from datetime import datetime, timedelta

    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=28.35, rightMargin=28.35,
                            topMargin=28.35, bottomMargin=28.35)
    styles = getSampleStyleSheet()
    word_wrap_style = styles["Normal"]
    word_wrap_style.wordWrap = 'CJK'
    pdf_content = []

    inventory_date = (datetime.strptime(forecastDate, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")

    orders = list(orders_collection.find({"delivery_date": forecastDate}))
    inventory = {i['sku']: i for i in Inventory.get_by_date(inventory_date).products}

    productos = {}
    for order in orders:
        for p in order['products']:
            sku = p['sku']
            if sku not in productos:
                productos[sku] = {
                    'name': p['name'],
                    'unit': p.get('unit', ''),
                    'cantidades': []
                }
            productos[sku]['cantidades'].append((f"{order['order_number']} - {order['customer_name']}", p['quantity']))

    pronostico = {}
    reactivo = {}

    for sku, p in sorted(productos.items(), key=lambda x: x[1]['name'].lower()):
        total_qty = sum(q for _, q in p['cantidades'])
        inv_qty = inventory.get(sku, {}).get('quantity', 0)
        destino = pronostico if inv_qty >= total_qty else reactivo

        destino.setdefault(p['name'], {'unit': p['unit'], 'rows': [], 'total': 0})
        destino[p['name']]['total'] += total_qty
        for cliente, qty in p['cantidades']:
            destino[p['name']]['rows'].append((cliente, qty))

    def build_table(titulo, data_dict):
        tabla = [[titulo]]
        tabla.append(['Producto (Total)', 'Cliente', 'Cantidad'])
        span_ranges = []

        for nombre, info in data_dict.items():
            total = info['total']
            unit = info['unit']
            filas = info['rows']
            producto_label = Paragraph(f"<b>{nombre} (Total: {total} {unit})</b>", word_wrap_style)
            start_row = len(tabla)

            for i, (cliente, qty) in enumerate(filas):
                tabla.append([
                    producto_label if i == 0 else '',  # solo en la primera fila
                    Paragraph(cliente, word_wrap_style),
                    Paragraph(f"{qty} {unit}", word_wrap_style)
                ])
            end_row = len(tabla) - 1
            span_ranges.append((0, start_row, 0, end_row))

        estilos = [
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#97D700')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.white),
            ('SPAN', (0, 0), (-1, 0)),  # título tabla
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('INNERGRID', (0, 1), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 1), (-1, -1), 0.5, colors.black),
        ]
        for r in span_ranges:
            estilos.append(('SPAN', r[0:2], r[2:4]))
            estilos.append(('VALIGN', r[0:2], r[2:4], 'MIDDLE'))

        t = Table(tabla, colWidths=[230, 200, 120])
        t.setStyle(TableStyle(estilos))
        return t

    if pronostico:
        pdf_content.append(build_table("PRONÓSTICO", pronostico))
        pdf_content.append(Spacer(1, 20))
    if reactivo:
        pdf_content.append(build_table("REACTIVO", reactivo))

    pdf.build(pdf_content)
    buffer.seek(0)
    return Response(buffer, mimetype='application/pdf',
                    headers={'Content-Disposition': f'inline; filename=picking_summary_{forecastDate}.pdf'})


# ═══════════════════════════════════════════════════════════════════════════════
#  REPORTE UE — EXCEL  (GET /api/reports/ue_excel/<startDate>/<endDate>)
# ═══════════════════════════════════════════════════════════════════════════════
@report_api.route('/ue_excel/<string:startDate>/<string:endDate>', methods=['GET'])
def get_ue_excel(startDate, endDate):
    """
    Genera reporte Excel de Unit Economics con 4 pestañas:
      - UE      : Dashboard con KPIs y fórmulas que referencian las otras pestañas
      - Pedidos : Órdenes agrupadas por día (GMV, AOV, ALV, métodos de pago…)
      - Compras : Detalle de compras con COGS por línea
      - Costos  : Costos operativos con resumen y detalle
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # ── PALETA DE COLORES ──────────────────────────────────────────────────────
    C_GREEN   = "97D700"   # Verde Frescapp
    C_DKGREEN = "2E7D32"   # Verde oscuro
    C_LTGREEN = "E8F5B0"   # Verde claro (filas alternas / highlights)
    C_GRAY    = "F5F5F5"   # Gris claro (filas alternas)
    C_WHITE   = "FFFFFF"

    # ── HELPERS DE ESTILO ──────────────────────────────────────────────────────
    def _fill(h):
        return PatternFill("solid", fgColor=h)

    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    _thin = Side(style="thin",   color="DDDDDD")
    _med  = Side(style="medium", color="97D700")

    def _border(bottom_med=False):
        b = _med if bottom_med else _thin
        return Border(left=_thin, right=_thin, top=_thin, bottom=b)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    FMT_COP = "#,##0"
    FMT_PCT = "0.00%"
    FMT_NUM = "#,##0"

    # ── NOMBRES DE MES EN ESPAÑOL ──────────────────────────────────────────────
    MES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
              7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

    start_dt = datetime.strptime(startDate, "%Y-%m-%d")
    end_dt   = datetime.strptime(endDate,   "%Y-%m-%d")

    months_in_range = set()
    cur = start_dt.replace(day=1)
    while cur <= end_dt:
        months_in_range.add(MES_ES[cur.month])
        cur = (cur + timedelta(days=32)).replace(day=1)

    # ── COLECCIONES ────────────────────────────────────────────────────────────
    _orders    = db["orders"]
    _purchases = db["purchases"]
    _costs     = db["costs"]
    _routes    = db["routes"]

    # ══════════════════════════════════════════════════════════════════════════
    # FETCH: PEDIDOS  (agrupado por día)
    # ══════════════════════════════════════════════════════════════════════════
    daily_orders  = {}
    all_customers = set()

    for orden in _orders.find({"delivery_date": {"$gte": startDate, "$lte": endDate}}, {
        "delivery_date": 1, "customer_email": 1, "products": 1
    }):
        fecha = orden.get("delivery_date", "")
        if fecha not in daily_orders:
            daily_orders[fecha] = dict(orders=0, lines=0, gmv=0.0,
                                       customers=set(),
                                       efectivo=0.0, davivienda=0.0,
                                       bancolombia=0.0, cartera=0.0)
        d = daily_orders[fecha]
        d["orders"] += 1
        email = orden.get("customer_email", "")
        d["customers"].add(email)
        all_customers.add(email)
        for p in orden.get("products", []):
            d["lines"] += 1
            d["gmv"] += float(p.get("price_sale", 0)) * float(p.get("quantity", 0))

    # Métodos de pago y logística desde rutas
    logistics_by_date = {}
    for ruta in _routes.find({"close_date": {"$gte": startDate, "$lte": endDate}}, {
        "close_date": 1, "cost": 1, "stops": 1
    }):
        fecha = ruta.get("close_date", "")
        logistics_by_date[fecha] = logistics_by_date.get(fecha, 0.0) + float(ruta.get("cost", 0))
        if fecha in daily_orders:
            d = daily_orders[fecha]
            for stop in ruta.get("stops", []):
                charged = float(stop.get("total_charged", 0))
                pm = stop.get("payment_method", "")
                if stop.get("status") == "Pagada":
                    if pm == "Davivienda":    d["davivienda"]  += charged
                    elif pm == "Bancolombia": d["bancolombia"] += charged
                    else:                     d["efectivo"]    += charged
                else:
                    d["cartera"] += charged

    total_logistics = sum(logistics_by_date.values())

    # ══════════════════════════════════════════════════════════════════════════
    # FETCH: COMPRAS  (detalle de productos)
    # ══════════════════════════════════════════════════════════════════════════
    compras_rows = []
    for pur in _purchases.find({"date": {"$gte": startDate, "$lte": endDate}}, {
        "date": 1, "purchase_number": 1, "proveedor": 1, "products": 1
    }):
        fecha   = pur.get("date", "")
        pur_num = pur.get("purchase_number", "")
        for p in pur.get("products", []):
            qty   = float(p.get("total_quantity_ordered") or p.get("total_quantity") or p.get("quantity") or 0)
            price = float(p.get("final_price_purchase") or p.get("price_purchase") or 0)
            compras_rows.append({
                "fecha": fecha, "purchase_number": pur_num,
                "proveedor": p.get("proveedor", pur.get("proveedor", "")),
                "sku":  p.get("sku",  ""),
                "name": p.get("name", ""),
                "quantity": qty, "unit": p.get("unit", ""),
                "price_purchase": price, "total": qty * price,
            })
    compras_rows.sort(key=lambda x: (x["fecha"], x["purchase_number"], x["name"]))

    # ══════════════════════════════════════════════════════════════════════════
    # FETCH: COSTOS
    # ══════════════════════════════════════════════════════════════════════════
    costos_summary = {
        "logistics":   total_logistics,
        "wh_rent":     0.0,
        "cost_tech":   0.0,
        "sales_force": 0.0,
        "cost_others": 0.0,
        "cost_supply": 0.0,
        "personnel":   0.0,
    }
    costs_detail = []

    q_diario  = {"typePeriod": "Diario",  "period": {"$gte": startDate, "$lte": endDate}}
    q_mensual = {"typePeriod": "Mensual", "period": {"$in": list(months_in_range)}}

    for cost in _costs.find({"$or": [q_diario, q_mensual]}, {
        "typeCost": 1, "amount": 1, "detail": 1, "period": 1, "typePeriod": 1
    }):
        tipo   = cost.get("typeCost", "")
        amount = float(cost.get("amount", 0))
        costs_detail.append({
            "tipo":         tipo,
            "detalle":      cost.get("detail", ""),
            "periodo":      cost.get("period", ""),
            "tipo_periodo": cost.get("typePeriod", ""),
            "monto":        amount,
        })
        if tipo in costos_summary:
            costos_summary[tipo] += amount

    # Agrega logística desde rutas como filas de detalle
    for fecha, cost_val in sorted(logistics_by_date.items()):
        costs_detail.append({
            "tipo": "logistics", "detalle": f"Costo logístico ruta {fecha}",
            "periodo": fecha, "tipo_periodo": "Diario", "monto": cost_val,
        })
    costs_detail.sort(key=lambda x: x["periodo"])

    # ══════════════════════════════════════════════════════════════════════════
    # CONSTRUCCIÓN DEL WORKBOOK
    # ══════════════════════════════════════════════════════════════════════════
    wb = Workbook()

    # ── HOJA: PEDIDOS ──────────────────────────────────────────────────────────
    ws_p = wb.active
    ws_p.title = "Pedidos"
    ws_p.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFGHIJK", [12, 10, 10, 16, 14, 14, 14, 14, 14, 14, 14]):
        ws_p.column_dimensions[col].width = w

    ws_p.row_dimensions[1].height = 40
    ws_p.merge_cells("A1:K1")
    c = ws_p["A1"]
    c.value = "PEDIDOS  —  FRESCAPP"
    c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 15)
    c.alignment = _align("center")

    ws_p.row_dimensions[2].height = 24
    ws_p.merge_cells("A2:K2")
    c = ws_p["A2"]
    c.value = f"Período:  {startDate}   →   {endDate}"
    c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 11)
    c.alignment = _align("center")

    ws_p.row_dimensions[3].height = 30
    P_HEADERS = ["Fecha", "# Órdenes", "# Líneas", "GMV", "AOV", "ALV",
                 "Clientes Únicos", "Efectivo", "Davivienda", "Bancolombia", "Cartera"]
    for ci, h in enumerate(P_HEADERS, 1):
        c = ws_p.cell(row=3, column=ci, value=h)
        c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 10)
        c.alignment = _align("center"); c.border = _border()
    ws_p.freeze_panes = "A4"

    P_START = 4
    sorted_dates = sorted(daily_orders)
    for ri, fecha in enumerate(sorted_dates, P_START):
        d  = daily_orders[fecha]
        bg = C_GRAY if ri % 2 == 0 else C_WHITE
        row_vals = [fecha, d["orders"], d["lines"], d["gmv"],
                    None, None,  # AOV, ALV — fórmulas abajo
                    len(d["customers"]),
                    d["efectivo"], d["davivienda"], d["bancolombia"], d["cartera"]]
        for ci, val in enumerate(row_vals, 1):
            c = ws_p.cell(row=ri, column=ci, value=val)
            c.fill = _fill(bg); c.font = _font(size=10)
            c.alignment = _align("center" if ci == 1 else "right")
            c.border = _border()
            if ci in (4, 8, 9, 10, 11):
                c.number_format = FMT_COP
        # AOV  =  GMV / Órdenes
        c = ws_p.cell(row=ri, column=5, value=f"=IF(B{ri}=0,0,D{ri}/B{ri})")
        c.fill = _fill(bg); c.font = _font(size=10)
        c.alignment = _align("right"); c.border = _border(); c.number_format = FMT_COP
        # ALV  =  GMV / Líneas
        c = ws_p.cell(row=ri, column=6, value=f"=IF(C{ri}=0,0,D{ri}/C{ri})")
        c.fill = _fill(bg); c.font = _font(size=10)
        c.alignment = _align("right"); c.border = _border(); c.number_format = FMT_COP

    P_LAST  = P_START + max(len(sorted_dates) - 1, 0)
    P_TOTAL = P_LAST + 1
    ws_p.row_dimensions[P_TOTAL].height = 28
    totals_p = [
        "TOTAL",
        f"=SUM(B{P_START}:B{P_LAST})",
        f"=SUM(C{P_START}:C{P_LAST})",
        f"=SUM(D{P_START}:D{P_LAST})",
        f"=IF(B{P_TOTAL}=0,0,D{P_TOTAL}/B{P_TOTAL})",
        f"=IF(C{P_TOTAL}=0,0,D{P_TOTAL}/C{P_TOTAL})",
        len(all_customers),                              # MUA período completo
        f"=SUM(H{P_START}:H{P_LAST})",
        f"=SUM(I{P_START}:I{P_LAST})",
        f"=SUM(J{P_START}:J{P_LAST})",
        f"=SUM(K{P_START}:K{P_LAST})",
    ]
    for ci, val in enumerate(totals_p, 1):
        c = ws_p.cell(row=P_TOTAL, column=ci, value=val)
        c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 10)
        c.alignment = _align("center" if ci == 1 else "right")
        c.border = _border(bottom_med=True)
        if ci in (4, 5, 6, 8, 9, 10, 11):
            c.number_format = FMT_COP

    # ── HOJA: COMPRAS ──────────────────────────────────────────────────────────
    ws_c = wb.create_sheet("Compras")
    ws_c.sheet_view.showGridLines = False

    for col, w in zip("ABCDEFGHI", [12, 14, 20, 12, 30, 10, 8, 16, 16]):
        ws_c.column_dimensions[col].width = w

    ws_c.row_dimensions[1].height = 40
    ws_c.merge_cells("A1:I1")
    c = ws_c["A1"]
    c.value = "COMPRAS  —  FRESCAPP"
    c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 15)
    c.alignment = _align("center")

    ws_c.row_dimensions[2].height = 24
    ws_c.merge_cells("A2:I2")
    c = ws_c["A2"]
    c.value = f"Período:  {startDate}   →   {endDate}"
    c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 11)
    c.alignment = _align("center")

    ws_c.row_dimensions[3].height = 30
    C_HEADERS = ["Fecha", "# Compra", "Proveedor", "SKU", "Producto",
                 "Cantidad", "Und", "Precio Compra", "Total"]
    for ci, h in enumerate(C_HEADERS, 1):
        c = ws_c.cell(row=3, column=ci, value=h)
        c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 10)
        c.alignment = _align("center"); c.border = _border()
    ws_c.freeze_panes = "A4"

    C_START = 4
    for ri, row in enumerate(compras_rows, C_START):
        bg = C_GRAY if ri % 2 == 0 else C_WHITE
        vals = [row["fecha"], row["purchase_number"], row["proveedor"],
                row["sku"], row["name"], row["quantity"], row["unit"],
                row["price_purchase"], row["total"]]
        for ci, val in enumerate(vals, 1):
            c = ws_c.cell(row=ri, column=ci, value=val)
            c.fill = _fill(bg); c.font = _font(size=10)
            c.alignment = _align("center" if ci in (1, 2, 4, 6, 7)
                                 else "left" if ci in (3, 5) else "right")
            c.border = _border()
            if ci in (8, 9):
                c.number_format = FMT_COP

    C_LAST  = C_START + max(len(compras_rows) - 1, 0)
    C_TOTAL = C_LAST + 1
    ws_c.row_dimensions[C_TOTAL].height = 28

    ws_c.merge_cells(f"A{C_TOTAL}:H{C_TOTAL}")
    c = ws_c.cell(row=C_TOTAL, column=1, value="TOTAL COMPRAS  (COGS)")
    c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 10)
    c.alignment = _align("right"); c.border = _border()

    cogs_formula = f"=SUM(I{C_START}:I{C_LAST})" if compras_rows else "=0"
    c = ws_c.cell(row=C_TOTAL, column=9, value=cogs_formula)
    c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 12)
    c.alignment = _align("right"); c.border = _border(bottom_med=True)
    c.number_format = FMT_COP

    # ── HOJA: COSTOS ───────────────────────────────────────────────────────────
    ws_k = wb.create_sheet("Costos")
    ws_k.sheet_view.showGridLines = False

    for col, w in zip("ABCDE", [24, 34, 15, 14, 18]):
        ws_k.column_dimensions[col].width = w

    ws_k.row_dimensions[1].height = 40
    ws_k.merge_cells("A1:E1")
    c = ws_k["A1"]
    c.value = "COSTOS OPERATIVOS  —  FRESCAPP"
    c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 15)
    c.alignment = _align("center")

    ws_k.row_dimensions[2].height = 24
    ws_k.merge_cells("A2:E2")
    c = ws_k["A2"]
    c.value = f"Período:  {startDate}   →   {endDate}"
    c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 11)
    c.alignment = _align("center")

    # Encabezado sección resumen (fila 3)
    ws_k.row_dimensions[3].height = 24
    ws_k.merge_cells("A3:E3")
    c = ws_k["A3"]
    c.value = "  RESUMEN POR CATEGORÍA"
    c.fill = _fill(C_LTGREEN); c.font = _font(True, C_DKGREEN, 11)
    c.alignment = _align("left")

    # Cabeceras del resumen (fila 4)
    ws_k.row_dimensions[4].height = 28
    for ci, h in enumerate(["Categoría", "Clave", "Total"], 1):
        c = ws_k.cell(row=4, column=ci, value=h)
        c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 10)
        c.alignment = _align("center"); c.border = _border()

    # Filas del resumen — posición fija (filas 5-11) para que la hoja UE
    # pueda referenciarlas con fórmulas estáticas (=Costos!C5, etc.)
    K_SUMM_START = 5
    SUMM_CATS = [
        ("Logística  (Last Mile)", "logistics",   costos_summary["logistics"]),
        ("Bodega  /  WH Rent",     "wh_rent",     costos_summary["wh_rent"]),
        ("Tecnología",             "cost_tech",   costos_summary["cost_tech"]),
        ("Fuerza de Ventas",       "sales_force", costos_summary["sales_force"]),
        ("Otros Costos",           "cost_others", costos_summary["cost_others"]),
        ("Suministros",            "cost_supply", costos_summary["cost_supply"]),
        ("Personal",               "personnel",   costos_summary["personnel"]),
    ]
    for i, (cat, typ, val) in enumerate(SUMM_CATS):
        rn = K_SUMM_START + i
        ws_k.row_dimensions[rn].height = 22
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        c = ws_k.cell(row=rn, column=1, value=cat)
        c.fill = _fill(bg); c.font = _font(size=10)
        c.alignment = _align("left"); c.border = _border()
        c = ws_k.cell(row=rn, column=2, value=typ)
        c.fill = _fill(bg); c.font = _font(size=9, color="888888")
        c.alignment = _align("center"); c.border = _border()
        c = ws_k.cell(row=rn, column=3, value=val)
        c.fill = _fill(bg); c.font = _font(size=10)
        c.alignment = _align("right"); c.border = _border()
        c.number_format = FMT_COP

    # Fila TOTAL OPEX (fila 12)
    K_OPEX_ROW = K_SUMM_START + len(SUMM_CATS)   # = 12
    ws_k.row_dimensions[K_OPEX_ROW].height = 30
    for ci, (val, fmt) in enumerate([
        ("TOTAL OPEX", None),
        ("",           None),
        (f"=SUM(C{K_SUMM_START}:C{K_OPEX_ROW-1})", FMT_COP),
    ], 1):
        c = ws_k.cell(row=K_OPEX_ROW, column=ci, value=val)
        c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 11)
        c.alignment = _align("right" if ci == 3 else "left")
        c.border = _border(bottom_med=True)
        if fmt:
            c.number_format = fmt

    # Separador visual
    ws_k.row_dimensions[K_OPEX_ROW + 1].height = 8

    # Encabezado sección detalle
    K_DET_HDR  = K_OPEX_ROW + 2
    K_DET_COLS = K_DET_HDR + 1
    ws_k.row_dimensions[K_DET_HDR].height = 24
    ws_k.merge_cells(f"A{K_DET_HDR}:E{K_DET_HDR}")
    c = ws_k.cell(row=K_DET_HDR, column=1, value="  DETALLE DE COSTOS")
    c.fill = _fill(C_LTGREEN); c.font = _font(True, C_DKGREEN, 11)
    c.alignment = _align("left")

    ws_k.row_dimensions[K_DET_COLS].height = 28
    for ci, h in enumerate(["Tipo", "Detalle", "Período", "Tipo Período", "Monto"], 1):
        c = ws_k.cell(row=K_DET_COLS, column=ci, value=h)
        c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 10)
        c.alignment = _align("center"); c.border = _border()
    ws_k.freeze_panes = f"A{K_DET_COLS}"

    K_DET_START = K_DET_COLS + 1
    for ri, row in enumerate(costs_detail, K_DET_START):
        bg = C_GRAY if ri % 2 == 0 else C_WHITE
        vals = [row["tipo"], row["detalle"], row["periodo"], row["tipo_periodo"], row["monto"]]
        for ci, val in enumerate(vals, 1):
            c = ws_k.cell(row=ri, column=ci, value=val)
            c.fill = _fill(bg); c.font = _font(size=10)
            c.alignment = _align("right" if ci == 5 else "left")
            c.border = _border()
            if ci == 5:
                c.number_format = FMT_COP

    # ── HOJA: UE (Unit Economics) — primera pestaña ────────────────────────────
    ws_u = wb.create_sheet("UE", 0)   # posición 0 = primera
    ws_u.sheet_view.showGridLines = False

    ws_u.column_dimensions["A"].width = 32
    ws_u.column_dimensions["B"].width = 24
    ws_u.column_dimensions["C"].width = 4    # columna separadora
    ws_u.column_dimensions["D"].width = 24
    ws_u.column_dimensions["E"].width = 18

    # Título principal
    ws_u.row_dimensions[1].height = 52
    ws_u.merge_cells("A1:E1")
    c = ws_u["A1"]
    c.value = "UNIT ECONOMICS  ·  FRESCAPP"
    c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 18)
    c.alignment = _align("center")

    # Período
    ws_u.row_dimensions[2].height = 28
    ws_u.merge_cells("A2:E2")
    c = ws_u["A2"]
    c.value = f"Período:  {startDate}   →   {endDate}"
    c.fill = _fill(C_GREEN); c.font = _font(True, C_WHITE, 12)
    c.alignment = _align("center")

    ws_u.row_dimensions[3].height = 12   # separador visual

    # ── Helpers para escribir la hoja UE ──────────────────────────────────────
    def ue_section(row, label):
        ws_u.row_dimensions[row].height = 26
        ws_u.merge_cells(f"A{row}:B{row}")
        c = ws_u.cell(row=row, column=1, value=f"  {label}")
        c.fill = _fill(C_LTGREEN); c.font = _font(True, C_DKGREEN, 11)
        c.alignment = _align("left")

    def ue_metric(row, label, formula, fmt=FMT_COP, bold=False, bg=C_WHITE):
        ws_u.row_dimensions[row].height = 24
        c = ws_u.cell(row=row, column=1, value=label)
        c.fill = _fill(bg); c.font = _font(bold, size=10)
        c.alignment = _align("left"); c.border = _border()
        c = ws_u.cell(row=row, column=2, value=formula)
        c.fill = _fill(bg); c.font = _font(bold, size=11)
        c.alignment = _align("right"); c.border = _border()
        if fmt:
            c.number_format = fmt

    def ue_total(row, label, formula, fmt=FMT_COP):
        ws_u.row_dimensions[row].height = 30
        c = ws_u.cell(row=row, column=1, value=label)
        c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 11)
        c.alignment = _align("left"); c.border = _border(bottom_med=True)
        c = ws_u.cell(row=row, column=2, value=formula)
        c.fill = _fill(C_DKGREEN); c.font = _font(True, C_WHITE, 12)
        c.alignment = _align("right"); c.border = _border(bottom_med=True)
        if fmt:
            c.number_format = fmt

    r = 4   # contador de fila

    # INGRESOS
    ue_section(r, "INGRESOS"); r += 1
    gmv_r = r
    ue_metric(r, "GMV  (Ventas Brutas)",
              f"=SUM(Pedidos!D{P_START}:D{P_LAST})", bg=C_GRAY); r += 1
    r += 1

    # COSTO DE BIENES VENDIDOS
    ue_section(r, "COSTO DE BIENES VENDIDOS  (COGS)"); r += 1
    cogs_r = r
    ue_metric(r, "COGS  (Costo de Bienes Vendidos)", cogs_formula, bg=C_GRAY); r += 1
    gp_r = r
    ue_metric(r, "Gross Profit  (GMV − COGS)",
              f"=B{gmv_r}-B{cogs_r}"); r += 1
    ue_metric(r, "Gross Profit  %",
              f"=IF(B{gmv_r}=0,0,B{gp_r}/B{gmv_r})", FMT_PCT, bg=C_GRAY); r += 1
    r += 1

    # COSTOS OPERATIVOS (OPEX)
    ue_section(r, "COSTOS OPERATIVOS  (OPEX)"); r += 1
    OPEX_UE = [
        ("Logística  (Last Mile)", f"=Costos!C{K_SUMM_START}"),
        ("Bodega  /  WH Rent",     f"=Costos!C{K_SUMM_START+1}"),
        ("Tecnología",             f"=Costos!C{K_SUMM_START+2}"),
        ("Fuerza de Ventas",       f"=Costos!C{K_SUMM_START+3}"),
        ("Otros Costos",           f"=Costos!C{K_SUMM_START+4}"),
        ("Suministros",            f"=Costos!C{K_SUMM_START+5}"),
        ("Personal",               f"=Costos!C{K_SUMM_START+6}"),
    ]
    for i, (lbl, fml) in enumerate(OPEX_UE):
        ue_metric(r, lbl, fml, bg=C_GRAY if i % 2 == 0 else C_WHITE); r += 1
    opex_total_r = r
    ue_total(r, "TOTAL OPEX", f"=Costos!C{K_OPEX_ROW}"); r += 1
    r += 1

    # RESULTADO
    ue_section(r, "RESULTADO"); r += 1
    net_r = r
    ue_metric(r, "Utilidad Neta  (Gross Profit − OPEX)",
              f"=B{gp_r}-B{opex_total_r}", bg=C_LTGREEN, bold=True); r += 1
    ue_metric(r, "Margen Neto  %",
              f"=IF(B{gmv_r}=0,0,B{net_r}/B{gmv_r})", FMT_PCT, bg=C_GRAY); r += 1
    r += 1

    # KPIs OPERACIONALES
    ue_section(r, "KPIs OPERACIONALES"); r += 1
    orders_r = r
    ue_metric(r, "# Órdenes",
              f"=SUM(Pedidos!B{P_START}:B{P_LAST})", FMT_NUM, bg=C_GRAY); r += 1
    lines_r = r
    ue_metric(r, "# Líneas",
              f"=SUM(Pedidos!C{P_START}:C{P_LAST})", FMT_NUM); r += 1
    ue_metric(r, "AOV  (Valor promedio por orden)",
              f"=IF(B{orders_r}=0,0,B{gmv_r}/B{orders_r})", FMT_COP, bg=C_GRAY); r += 1
    ue_metric(r, "ALV  (Valor promedio por línea)",
              f"=IF(B{lines_r}=0,0,B{gmv_r}/B{lines_r})", FMT_COP); r += 1
    ue_metric(r, "MUA  (Clientes únicos del período)",
              f"=Pedidos!G{P_TOTAL}", FMT_NUM, bg=C_GRAY); r += 1

    # Panel derecho — Métodos de pago (columnas D-E)
    def ue_right(row, label, formula, fmt=FMT_COP, bg=C_WHITE):
        c = ws_u.cell(row=row, column=4, value=label)
        c.fill = _fill(bg); c.font = _font(size=10)
        c.alignment = _align("left"); c.border = _border()
        c = ws_u.cell(row=row, column=5, value=formula)
        c.fill = _fill(bg); c.font = _font(size=11)
        c.alignment = _align("right"); c.border = _border()
        if fmt:
            c.number_format = fmt

    # Encabezado panel derecho en fila 4
    ws_u.merge_cells("D4:E4")
    c = ws_u["D4"]
    c.value = "  MÉTODOS DE PAGO"
    c.fill = _fill(C_LTGREEN); c.font = _font(True, C_DKGREEN, 11)
    c.alignment = _align("left")

    ue_right(5, "Efectivo",    f"=SUM(Pedidos!H{P_START}:H{P_LAST})", bg=C_GRAY)
    ue_right(6, "Davivienda",  f"=SUM(Pedidos!I{P_START}:I{P_LAST})")
    ue_right(7, "Bancolombia", f"=SUM(Pedidos!J{P_START}:J{P_LAST})", bg=C_GRAY)
    ue_right(8, "Cartera",     f"=SUM(Pedidos!K{P_START}:K{P_LAST})")

    # ── SERIALIZAR Y RETORNAR ──────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"UE_Frescapp_{startDate}_{endDate}.xlsx"
    resp = Response(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp
